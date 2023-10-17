import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import psycopg2
import datetime
import os
from tkinter import filedialog
from openpyxl import Workbook


cursor = None

def id():
    cur = con.cursor()
    lastId = "SELECT MAX(pwd_id) FROM passwords"
    cur.execute(lastId)
    result = cur.fetchone()
    if result:
        max_pwd_id = result[0]
        with open("pwd_id.txt", 'w') as file:
            file.write(str(max_pwd_id))
        print(f"Maksymalne pwd_id zapisane do pliku")
    else:
        print("Brak rekordów w tabeli 'passwords'")

#Funkcja wyświetlania bazy danych
def view():
    try:
        cur1 = con.cursor()
        cur1.execute("SELECT * FROM passwords")
        rows = cur1.fetchall()
        messagebox.showinfo("Info", "Pomyślnie pobrano dane!")
    except:
        messagebox.showerror("Błąd", "Nie udało się pobrać danych")

    for row in rows:
        print(row)
        tree.insert("", tk.END, values=row)


#Funkcja akutalizowania
def update():
    if os.path.isfile("pwd_id.txt"):
        try:
            with open("pwd_id.txt") as file:
                lastId = int(file.readline())
                print("LAST RECORD: ", lastId)
        except ValueError:
            # Jeśli nie można sparsować daty, ustaw początkową datę na bardzo dawno temu
            lastId = 50
    else:
        lastId = 50

    try:
        con
        cur2 = con.cursor()
        query = "SELECT * FROM passwords WHERE pwd_id > %s"
        cur2.execute(query, (lastId,))
        result = cur2.fetchall()
        print("Query: ", query)
        print("Result: ", result)

        if result:
            for row in result:
                print(row)
                tree.insert("", tk.END, values=row)

            # Zaktualizuj plik pwd_id.txt z nowym najwyższym pwd_id
            new_last_id = max([row[0] for row in result])
            with open("pwd_id.txt", "w") as file:
                file.write(str(new_last_id))

    except Exception as e:
        print("Błąd bazy danych:", e)

#Zamykanie okna logowania, otwieranie okna aplikacji
def showMainWindow(loginWindow):
    if loginWindow.winfo_exists():
        loginWindow.destroy()
    passwords.deiconify()



#Parametry do połączenia z bazą danych
con = psycopg2.connect(
        host = "localhost",
        port = "5432",
        database = "Passwords",
        user = "postgres",
        password = "Logintech"
    )
#Definiowanie funkcji logowania
def login():
    f=open("pwd.txt","r")
    lines=f.readlines()
    username = "Admin"
    password = lines[0]
    f.close()
    lastStart = datetime.datetime.now()
    f2 = open("ostatnie_uruchomienie.txt", "w")
    f2.write(lastStart.strftime("%Y-%m-%d %H:%M:%S"))
    f2.close()
    global cursor
    if userEntry.get() == username and userPass.get() == password:
        con,
        cursor = con.cursor()
        messagebox.showinfo("Sukces", "Pomyślnie zalogowano do bazy danych")
        showMainWindow(loginWindow)
        view()
        id()
    elif userEntry.get() == username and userPass.get() != password:
        messagebox.showerror("Błąd", "Wprowadzono błędne hasło")
    elif userEntry.get() != username and userPass.get() == password:
        messagebox.showerror("Błąd", "Wprowadzono błędną nazwę użytkownika")
    else:
        messagebox.showerror("Błąd", "Wprowadzono błędne hasło lub nazwę użytkownika")


#Tworzenie okna logowania
loginWindow = tk.Tk()
loginWindow.geometry("400x400")
loginWindow.title("Zaloguj się do bazy danych")

loginWindowLabel = tk.Label(loginWindow, text="Okno logowania")
loginWindowLabel.pack(pady=20)

loginFrame = tk.Frame(loginWindow)
loginFrame.pack(pady=20, padx=40, fill="both", expand=True)

loginFrameLabel = tk.Label(loginFrame, text="Zaloguj się")
loginFrameLabel.pack(pady=12, padx=10)

userEntry = tk.Entry(loginFrame)
userEntry.pack(pady=12, padx=10)

userPass = tk.Entry(loginFrame, show="*")
userPass.pack(pady=12, padx=10)

loginButton = tk.Button(loginFrame, text="Zaloguj się", command=login)
loginButton.pack(pady=12, padx=10)


#Funkcja tworzenia nowego wiersza
def addNewRow():
    for _ in range(3): #Tworzenie 3 pól entry w jednym wierszu
        newField = tk.Entry(frameInsideCanvas, width=20) #Tworzenie nowego pola entry w ramce, o szer 20
        newField.grid(column=len(entryField) % 3, row=len(entryField) // 3 + 2, padx=10, pady=5) #Umieszczenie nowego pola Entry
        entryField.append(newField) #Dodawanie nowych pól do listy entryField

#Funkcja usuwania ostatniego wiersza
def delLastRow():
    for _ in range(3):
        if entryField:
            #Usuwanie ostatnich pól entry
            entryField[-1].destroy()
            entryField.pop()
    updateCanvasScrollregion() #Aktualizacja wielkości scrollbara

#Aktualizacja wielkości scrolbara
def updateCanvasScrollregion():
    bbox = canvas.bbox("all") #Pobranie obszaru ograniczającego w canvasie
    canvas.config(scrollregion=bbox) #Aktualizacja obszaru dla scrollbara


#Zamykanie programu
def closing():
    passwords.destroy()
    loginWindow.destroy()


#Lista pól Entry
entryField = []

#Tworzenie okienka
passwords = tk.Tk()
passwords.title("Hasła")
passwords.geometry("1050x480")
passwords.withdraw()

#Tworzenie zakładek
tab = ttk.Notebook(passwords)
tab.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
tab1 = ttk.Frame(tab)
tab2 = ttk.Frame(tab)
tab.add(tab1, text="Dodawanie hasła")
tab.add(tab2, text="Zapisane hasła")

#Tworzenie przestrzeni dla Canvas
frameForCanvas = tk.Frame(tab1, relief="sunken", highlightthickness=0, background="white" )
frameForCanvas.grid(row=0, column=0, sticky="nsew")

#Widget Canvas
canvas = tk.Canvas(frameForCanvas, height=350, width=750, highlightthickness=0, relief="sunken")
canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

#Pasek przewijania
scrollbar = tk.Scrollbar(frameForCanvas, orient=tk.VERTICAL, command=canvas.yview)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
canvas.configure(yscrollcommand=scrollbar.set)

#Pole robocze
frameInsideCanvas = tk.Frame(canvas)
canvas.create_window((0, 0), window=frameInsideCanvas, anchor=tk.NW)

#Zakładka 1 - tworzenie pól do dodawania haseł
serviceLabel = tk.Label(frameInsideCanvas, text="Usługa")
serviceLabel.grid(column=0, row=0, padx=10, pady=5)
serviceEntry = tk.Entry(frameInsideCanvas, width=20)
serviceEntry.grid(column = 0, row = 1, padx=10, pady=5)

loginLabel = tk.Label(frameInsideCanvas, text="Login")
loginLabel.grid(column=1, row=0)
loginEntry = tk.Entry(frameInsideCanvas, width=20)
loginEntry.grid(column=1, row=1,padx=10, pady=5)

passwordLabel = tk.Label(frameInsideCanvas, text="Hasło")
passwordLabel.grid(column=2, row=0)
passwordEntry = tk.Entry(frameInsideCanvas, width=20)
passwordEntry.grid(column=2, row=1, padx=10, pady=5)

#Przycisk do dodawania pól
addFieldButton = tk.Button(frameInsideCanvas, text=" + ", command=lambda: [addNewRow(), updateCanvasScrollregion()])
addFieldButton.grid(column=3, row=1, padx=8, pady=5)

#Przycisk do usuwania ostatniego pola
undoButton = tk.Button(frameInsideCanvas, text=" - ", command=delLastRow)
undoButton.grid(column=4, row=1, padx=8, pady=5)

#Blokada scrollbara przed dodaniem kilku pól
scrollbar.config(command=canvas.yview)

#Dodanie trzech początkowych pól Entry do listy
entryField.extend([serviceEntry, loginEntry, passwordEntry])

def addToDatabase():
    #Pobranie danych z pól Entry
    rowsToInsert = []

    for i in range(len(entryField)//3):
        service = entryField[i*3].get()
        login = entryField[i*3+1].get()
        password = entryField[i*3+2].get()

        #Sprawdzenie czy wszystkie komórki w rzędzie są wypełnione
        if service and login and password:
            rowsToInsert.append((service, login, password))

        #Zapytanie do SQL
    if rowsToInsert:
            insertQuery = "INSERT INTO passwords (service, login, password) VALUES (%s, %s, %s)"

            try:
                #Wykonanie zapytania
                cursor = con.cursor()
                cursor.executemany(insertQuery, rowsToInsert)
                con.commit()
                lastRecord = datetime.datetime.now()
                f = open("ostatni_rekord.txt", "w")
                f.write(lastRecord.strftime("%Y-%m-%d %H:%M:%S"))
                f.close()
                messagebox.showinfo("Sukces", "Dane zostały dodane do bazy danych")

                #Wyczyszczenie pól entry
                for entry in entryField:
                    entry.delete(0, tk.END)
            except Exception as e:
                print(f"Błąd: {str(e)}")
                messagebox.showerror("Błąd", "Podczas dodawania danych do bazy dany wystąpił błąd. "
                                             "Sprawdź połączenienie z bazą")
    else:
        messagebox.showerror("Błąd", "Wprowadź wszystkie dane przed dodaniem do bazy danych")

#Przycisk do dodania do bazy danych
addPgButton = tk.Button(tab1, text="Dodaj do bazy danych", command=addToDatabase)
addPgButton.grid(column=0, row=17)

#Tworzenie 2 zakładki


#Tworzenie widoku Treeview
tree = ttk.Treeview(tab2, column=("c1", "c2", "c3", "c4", "c5"), show='headings')
tree.column("#1", anchor=tk.CENTER)
tree.heading("#1", text="ID")
tree.column("#2", anchor=tk.CENTER)
tree.heading("#2", text="Usługa")
tree.column("#3", anchor=tk.CENTER)
tree.heading("#3", text="Login")
tree.column("#4", anchor=tk.CENTER)
tree.heading("#4", text="Hasło")
tree.column("#5", anchor=tk.CENTER)
tree.heading("#5", text="Data dodania")
tree.pack()
db_button= tk.Button(tab2, text="Zaktualizuj dane", command=update)
db_button.pack(pady=10, side=tk.TOP, anchor="n")


#Tworzenie funkcji sortowania
def sortRecords(columnName):
    items = list(tree.get_children(''))

    items.sort(key=lambda item: tree.set(item, columnName))

    sortedItems = [tree.item(item, 'values') for item in items]

    tree.delete(*tree.get_children())

    for values in sortedItems:
        tree.insert("", "end", values=values)


sortList = ttk.Combobox(tab2)
sortList["values"] = ["#1", "#2", "#3", "#4", "#5"]
sortList.current(0)
sortList.pack(side=tk.LEFT, anchor="n", padx=10, pady=10)

sort_button = tk.Button(tab2, text="Sortuj rosnąco", command=lambda: sortRecords(sortList.get()))
sort_button.pack(side=tk.LEFT, anchor="n", padx=10, pady=10)

#Sortowanie malejące
def sortRecordReverse(columnName):
    items = list(tree.get_children())

    items.sort(key=lambda item: tree.set(item, columnName))

    for columnName in ["#1", "#2", "#3", "#4", "#5"]:
        items.reverse()

    sortedItems = [tree.item(item, 'values') for item in items]

    tree.delete(*tree.get_children())

    for values in sortedItems:
        tree.insert("", "end", values=values)

sortReverseButton = tk.Button(tab2, text="Sortuj malejąco", command=lambda: sortRecordReverse(sortList.get()))
sortReverseButton.pack()


#Funkcja wyszukaj
def search():
    query = searchEntry.get()
    selections = []
    for child in tree.get_children():
        if query in tree.item(child)["values"]:
            print(tree.item(child)["values"])
            selections.append(child)
    print("Wyszukiwanie zakończone")
    tree.selection_set(selections)

values = []
searchLabel = tk.Label(tab2, text="Search:")
searchLabel.pack()
searchEntry = tk.Entry(tab2, width=15)
searchEntry.pack()
searchButton = tk.Button(tab2, text="search", width=10, command=search)
searchButton.pack()

#Funkcja exportowania
def export(treeview):
    filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
    if filename:
        workbook = Workbook()
        worksheet = workbook.active

        #Pobranie nazw kolumn z treeview
        columns = [treeview.heading(column)["text"] for column in treeview["columns"]]

        #Nagłówki kolumn w arkuszu
        for col_idx, column_title in enumerate(columns):
            worksheet.cell(row = 1, column=col_idx + 1, value = column_title)

        for row_idx, item in enumerate(treeview.get_children()):
            for col_idx, column in enumerate(columns):
                value = treeview.item(item, "values")[col_idx]
                worksheet.cell(row=row_idx + 2, column=col_idx + 1, value=value)

        workbook.save(filename)
        print(f"Exported to: {filename}")

export_button = tk.Button(tab2, text="Export to Excel", command=lambda: export(tree))
export_button.pack()


#Funkcja usuwania
def delete_selected_record():
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showwarning("Błąd", "Proszę zaznaczyć rekord do usunięcia.")
        return

    for item in selected_item:
        # Pobierz ID zaznaczonego rekordu
        record_id = tree.item(item, "values")[0]

        # Usuń rekord z bazy danych
        try:
            cursor.execute("DELETE FROM passwords WHERE pwd_id = %s", (record_id,))
            con.commit()
            tree.delete(item)  # Usuń zaznaczony rekord z widoku Treeview
        except Exception as e:
            messagebox.showerror("Błąd", f"Błąd podczas usuwania rekordu: {str(e)}")

    messagebox.showinfo("Sukces", "Rekordy zostały usunięte z bazy danych.")

# Tworzenie przycisku do usuwania zaznaczonego rekordu
delete_button = tk.Button(tab2, text="Usuń rekord", command=delete_selected_record)
delete_button.pack()


if __name__ == "__main__":
    loginWindow.protocol("WM_DELETE_WINDOW", closing)
    loginWindow.mainloop()
